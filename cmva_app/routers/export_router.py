from fastapi import APIRouter, Request, Depends
from fastapi.responses import StreamingResponse, RedirectResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io
from database import get_db
import auth, models, utils
from routers.registros_router import _build_query

router = APIRouter(prefix="/exportar", tags=["exportacion"])


def _get_user(request, db):
    user = auth.get_current_user(request, db)
    if not user:
        return None, RedirectResponse("/login", status_code=302)
    return user, None


def _get_registros(db, params):
    return _build_query(db, params).order_by(models.RegistroRayosX.fecha_hora_toma.asc()).all()


_ESTUDIO_NOMBRE = {"RX": "Rayos X", "TAC": "TAC", "ECO": "Ecografía"}


def _row_data(r):
    # Orden exacto según PLANTILLA.xlsx + VOLUMEN_CONTRASTE_ML después de mAs
    # LECTURA_RADIOLOGO = nombre del radiólogo (persona)
    # OBSERVACION_RADIOLOGO = texto del informe/lectura
    return [
        _ESTUDIO_NOMBRE.get(r.tipo_estudio_principal or "RX", r.tipo_estudio_principal or "Rayos X"),  # ESTUDIO
        r.codigo_registro,                          # ID_REGISTRO
        utils.format_dt(r.fecha_hora_toma),         # FECHA_HORA_TOMA
        utils.format_dt(r.fecha_hora_orden),        # FECHA_HORA_ORDEN
        r.identificacion,                           # IDENTIFICACION
        r.nombre_paciente,                          # NOMBRE_Y_APELLIDO
        r.contrastado_si_no or "NO",               # CONTRASTADO_SI_NO
        r.estudio.nombre if r.estudio else "",      # TIPO_ESTUDIO
        r.codigo_tac or "",                         # CODIGO_TAC
        r.servicio.nombre if r.servicio else "",    # SERVICIO
        r.profesional.nombre if r.profesional else "",  # PROFESIONAL
        r.numero_tomas,                             # NUMERO_TOMAS
        r.kv,                                       # KV
        r.mas,                                      # mAs
        r.volumen_contraste_ml if r.volumen_contraste_ml is not None else "",  # VOLUMEN_CONTRASTE_ML
        r.observaciones_tecnologo or "",            # OBSERVACIONES_TECNOLOGO
        r.tecnologo.nombre if r.tecnologo else "",  # TECNOLOGO
        r.reporte.nombre if r.reporte else "",      # REPORTE
        "SI" if r.rechazada else "NO",              # RECHAZADA
        r.causa_rechazo.descripcion if r.causa_rechazo else "",  # CAUSA_RECHAZO
        utils.format_dt(r.fecha_hora_lectura_radiologo),         # FECHA_HORA_LECTURA_RADIOLOGO
        r.radiologo.nombre if r.radiologo else "",  # LECTURA_RADIOLOGO = nombre del radiólogo
        r.transcriptora.nombre if r.transcriptora else "",       # TRANSCRIPTORA
        r.lectura_radiologo or r.observaciones_radiologo or r.observaciones or "",  # OBSERVACION_RADIOLOGO = texto
        r.estado,                                   # ESTADO
    ]


HEADERS = [
    "ESTUDIO", "ID_REGISTRO", "FECHA_HORA_TOMA", "FECHA_HORA_ORDEN",
    "IDENTIFICACION", "NOMBRE_Y_APELLIDO", "CONTRASTADO_SI_NO", "TIPO_ESTUDIO",
    "CODIGO_TAC", "SERVICIO", "PROFESIONAL", "NUMERO_TOMAS", "KV", "mAs",
    "VOLUMEN_CONTRASTE_ML", "OBSERVACIONES_TECNOLOGO", "TECNOLOGO", "REPORTE",
    "RECHAZADA", "CAUSA_RECHAZO", "FECHA_HORA_LECTURA_RADIOLOGO", "LECTURA_RADIOLOGO",
    "TRANSCRIPTORA", "OBSERVACION_RADIOLOGO", "ESTADO",
]


def _parse_params(kwargs):
    return {k: v for k, v in kwargs.items() if v}


@router.get("/excel")
async def export_excel(
    request: Request,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo_estudio_principal: Optional[str] = None,
    tecnologo_id: Optional[str] = None,
    radiologo_id: Optional[str] = None,
    servicio_id: Optional[str] = None,
    estudio_id: Optional[str] = None,
    estado: Optional[str] = None,
    rechazada: Optional[str] = None,
    identificacion: Optional[str] = None,
    nombre_paciente: Optional[str] = None,
    db: Session = Depends(get_db),
):
    user, redir = _get_user(request, db)
    if redir:
        return redir
    if not utils.can_export(user):
        return RedirectResponse("/registros", status_code=302)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return StreamingResponse(iter(["openpyxl no instalado"]), media_type="text/plain")

    params = _parse_params({
        "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
        "tipo_estudio_principal": tipo_estudio_principal,
        "tecnologo_id": tecnologo_id, "radiologo_id": radiologo_id,
        "servicio_id": servicio_id, "estudio_id": estudio_id,
        "estado": estado, "rechazada": rechazada,
        "identificacion": identificacion, "nombre_paciente": nombre_paciente,
    })

    registros = _get_registros(db, params)
    wb = Workbook()
    ws = wb.active
    ws.title = "Imagenes Diagnosticas"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, r in enumerate(registros, 2):
        for col_num, value in enumerate(_row_data(r), 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"CMVA_ImagenesDiagnosticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/txt")
async def export_txt(
    request: Request,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo_estudio_principal: Optional[str] = None,
    tecnologo_id: Optional[str] = None,
    radiologo_id: Optional[str] = None,
    servicio_id: Optional[str] = None,
    estudio_id: Optional[str] = None,
    estado: Optional[str] = None,
    rechazada: Optional[str] = None,
    identificacion: Optional[str] = None,
    nombre_paciente: Optional[str] = None,
    db: Session = Depends(get_db),
):
    user, redir = _get_user(request, db)
    if redir:
        return redir
    if not utils.can_export(user):
        return RedirectResponse("/registros", status_code=302)

    params = _parse_params({
        "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
        "tipo_estudio_principal": tipo_estudio_principal,
        "tecnologo_id": tecnologo_id, "radiologo_id": radiologo_id,
        "servicio_id": servicio_id, "estudio_id": estudio_id,
        "estado": estado, "rechazada": rechazada,
        "identificacion": identificacion, "nombre_paciente": nombre_paciente,
    })

    registros = _get_registros(db, params)

    def generate():
        yield "|".join(HEADERS) + "\n"
        for r in registros:
            yield "|".join(str(v) for v in _row_data(r)) + "\n"

    filename = f"CMVA_ImagenesDiagnosticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    return StreamingResponse(
        generate(),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
