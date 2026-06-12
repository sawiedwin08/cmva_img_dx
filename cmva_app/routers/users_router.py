from fastapi import APIRouter, Request, Form, Depends, HTTPException, Response, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io
from database import get_db
import auth, models, utils
from templates_config import templates

router = APIRouter(prefix="/admin", tags=["usuarios"])


def _require_admin(request, db):
    user = auth.get_current_user(request, db)
    if not user:
        return None, RedirectResponse("/login", status_code=302)
    if not utils.is_admin(user):
        return None, RedirectResponse("/registros", status_code=302)
    return user, None


@router.get("/usuarios", response_class=HTMLResponse)
async def list_users(request: Request, response: Response, db: Session = Depends(get_db)):
    user, redir = _require_admin(request, db)
    if redir:
        return redir
    usuarios = db.query(models.Usuario).order_by(models.Usuario.nombre_completo).all()
    roles = db.query(models.Rol).filter_by(estado=True).all()
    especialidades = db.query(models.Especialidad).filter_by(estado=True).all()
    flash = utils.get_flash(request, response)
    return templates.TemplateResponse("admin/usuarios.html", {
        "request": request, "user": user, "flash": flash,
        "usuarios": usuarios, "roles": roles, "especialidades": especialidades,
    })


@router.post("/usuarios/nuevo")
async def create_user(
    request: Request,
    nombre_completo: str = Form(...),
    usuario_login: str = Form(...),
    correo: Optional[str] = Form(None),
    password: str = Form(...),
    rol_id: int = Form(...),
    tipo_documento: Optional[str] = Form(None),
    numero_documento: Optional[str] = Form(None),
    especialidad_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
):
    user, redir = _require_admin(request, db)
    if redir:
        return redir
    redir_back = RedirectResponse("/admin/usuarios", status_code=303)
    if db.query(models.Usuario).filter_by(usuario=usuario_login).first():
        utils.set_flash(redir_back, "error", f"El usuario '{usuario_login}' ya existe.")
        return redir_back
    nuevo = models.Usuario(
        nombre_completo=nombre_completo.strip().upper(),
        usuario=usuario_login.strip().lower(),
        correo=correo or None,
        password_hash=auth.hash_password(password),
        rol_id=rol_id,
        tipo_documento=tipo_documento or None,
        numero_documento=numero_documento or None,
        especialidad_id=especialidad_id or None,
        estado=True,
    )
    db.add(nuevo)
    db.flush()
    rol = db.query(models.Rol).filter_by(id=rol_id).first()
    if rol and rol.nombre == "TECNÓLOGO":
        db.add(models.Tecnologo(nombre=nombre_completo.strip().upper(), usuario_id=nuevo.id))
    elif rol and rol.nombre == "RADIÓLOGO":
        db.add(models.Radiologo(nombre=nombre_completo.strip().upper(), usuario_id=nuevo.id, especialidad_id=especialidad_id))
    db.commit()
    utils.set_flash(redir_back, "success", "Usuario creado correctamente.")
    return redir_back


@router.post("/usuarios/{user_id}/editar")
async def edit_user(
    request: Request,
    user_id: int,
    nombre_completo: str = Form(...),
    correo: Optional[str] = Form(None),
    rol_id: int = Form(...),
    tipo_documento: Optional[str] = Form(None),
    numero_documento: Optional[str] = Form(None),
    especialidad_id: Optional[int] = Form(None),
    nuevo_password: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    admin_user, redir = _require_admin(request, db)
    if redir:
        return redir
    target = db.query(models.Usuario).filter_by(id=user_id).first()
    if not target:
        raise HTTPException(status_code=404)
    target.nombre_completo = nombre_completo.strip().upper()
    target.correo = correo or None
    target.rol_id = rol_id
    target.tipo_documento = tipo_documento or None
    target.numero_documento = numero_documento or None
    target.especialidad_id = especialidad_id or None
    if nuevo_password:
        target.password_hash = auth.hash_password(nuevo_password)
    db.commit()
    redir_ok = RedirectResponse("/admin/usuarios", status_code=303)
    utils.set_flash(redir_ok, "success", "Usuario actualizado.")
    return redir_ok


@router.post("/usuarios/{user_id}/toggle")
async def toggle_user(request: Request, user_id: int, db: Session = Depends(get_db)):
    admin_user, redir = _require_admin(request, db)
    if redir:
        return redir
    target = db.query(models.Usuario).filter_by(id=user_id).first()
    if not target:
        raise HTTPException(status_code=404)
    redir_back = RedirectResponse("/admin/usuarios", status_code=303)
    if target.usuario in ("admin", "ADM"):
        utils.set_flash(redir_back, "error", "No se puede desactivar el usuario administrador principal.")
        return redir_back
    target.estado = not target.estado
    db.commit()
    estado_txt = "activado" if target.estado else "inactivado"
    utils.set_flash(redir_back, "success", f"Usuario {estado_txt} correctamente.")
    return redir_back


@router.get("/usuarios/template-excel")
async def download_user_template(request: Request, db: Session = Depends(get_db)):
    admin_user, redir = _require_admin(request, db)
    if redir:
        return redir
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    headers = ["TIPO_DOCUMENTO", "NUMERO_DOCUMENTO", "NOMBRE_COMPLETO"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 25
    # Fila de ejemplo
    ws.append(["CC", "12345678", "EJEMPLO APELLIDO NOMBRE"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_usuarios.xlsx"},
    )


@router.post("/usuarios/importar-excel")
async def importar_usuarios_excel(
    request: Request,
    rol_id: int = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    admin_user, redir = _require_admin(request, db)
    if redir:
        return redir

    redir_back = RedirectResponse("/admin/usuarios", status_code=303)
    rol = db.query(models.Rol).filter_by(id=rol_id).first()
    if not rol:
        utils.set_flash(redir_back, "error", "Rol no encontrado.")
        return redir_back

    try:
        from openpyxl import load_workbook
        contenido = await archivo.read()
        wb = load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as e:
        utils.set_flash(redir_back, "error", f"No se pudo leer el archivo: {e}")
        return redir_back

    creados = 0
    omitidos = 0
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tipo_doc = str(row[0]).strip().upper() if row[0] else ""
        num_doc = str(row[1]).strip() if row[1] else ""
        nombre = str(row[2]).strip().upper() if row[2] else ""

        if not num_doc or not nombre:
            omitidos += 1
            continue

        # Usuario = número de documento (lowercase)
        usuario_login = num_doc.lower()
        if db.query(models.Usuario).filter_by(usuario=usuario_login).first():
            omitidos += 1
            continue

        try:
            nuevo = models.Usuario(
                nombre_completo=nombre,
                usuario=usuario_login,
                tipo_documento=tipo_doc or None,
                numero_documento=num_doc,
                password_hash=auth.hash_password(num_doc),
                rol_id=rol_id,
                estado=True,
            )
            db.add(nuevo)
            db.flush()
            if rol.nombre == "TECNÓLOGO":
                db.add(models.Tecnologo(nombre=nombre, usuario_id=nuevo.id))
            elif rol.nombre == "RADIÓLOGO":
                db.add(models.Radiologo(nombre=nombre, usuario_id=nuevo.id))
            db.commit()
            creados += 1
        except Exception as e:
            db.rollback()
            errores.append(f"Fila {row_num}: {str(e)[:60]}")

    msg = f"Importación completada: {creados} usuario(s) creado(s), {omitidos} omitido(s)."
    if errores:
        msg += f" Errores: {len(errores)}."
    utils.set_flash(redir_back, "success" if not errores else "warning", msg)
    return redir_back
