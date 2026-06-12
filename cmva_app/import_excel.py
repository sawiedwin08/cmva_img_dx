"""
Script de importación histórica del archivo Excel de CMVA.
Ejecutar UNA sola vez desde la raíz del proyecto cmva_app:
  python import_excel.py <ruta_del_excel>

Ejemplo:
  python import_excel.py "../Informe de rayos X CMVA.xlsx"
"""
import sys
import re
from datetime import datetime
from pathlib import Path
from database import SessionLocal, engine, Base
from seed import seed_initial_data
import models

Base.metadata.create_all(bind=engine)

# Normalización de nombres de estudio (variaciones → nombre canónico)
ESTUDIO_ALIASES = {
    "RX ABD": "RX ABDOMEN",
    "RX  C CERVICAL": "RX COLUMNA CERVICAL",
    "RX C CERVICAL": "RX COLUMNA CERVICAL",
    "RX C. CERVICAL": "RX COLUMNA CERVICAL",
    "RX C.CERVICAL": "RX COLUMNA CERVICAL",
    "RX COL CERVICAL": "RX COLUMNA CERVICAL",
    "RX C DORSAL": "RX COLUMNA DORSAL",
    "RX C TORACICA": "RX COLUMNA DORSAL",
    "RX C DORSOLUMBAR": "RX COLUMNA DORSOLUMBAR",
    "RX CDL": "RX COLUMNA DORSOLUMBAR",
    "RX CLS": "RX COLUMNA LUMBAR",
    "RX  CLS": "RX COLUMNA LUMBAR",
    "RX  MANO": "RX MANO",
    "RX  RODILLA": "RX RODILLA",
    "RC CADERA": "RX CADERA",
}

# Normalización de servicios
SERVICIO_ALIASES = {
    "C EXT": "CONSULTA EXTERNA",
    "CEXT": "CONSULTA EXTERNA",
    "C.EXT": "CONSULTA EXTERNA",
    "C. EXT": "CONSULTA EXTERNA",
    "URG": "URGENCIAS",
    "HOSP": "HOSPITALIZACIÓN",
    "UCI": "UCI",
    "QX": "QUIRÓFANO",
    "PQX": "PREQUIRÚRGICO",
}


def normalize_name(s):
    if not s:
        return ""
    s = str(s).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s


def get_or_create_estudio(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    nombre = ESTUDIO_ALIASES.get(nombre, nombre)
    if not nombre:
        return None
    estudio = db.query(models.Estudio).filter_by(nombre=nombre).first()
    if not estudio:
        estudio = models.Estudio(nombre=nombre, estado=True)
        db.add(estudio)
        db.flush()
    return estudio


def get_or_create_servicio(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    nombre = SERVICIO_ALIASES.get(nombre, nombre)
    if not nombre:
        return None
    servicio = db.query(models.Servicio).filter_by(nombre=nombre).first()
    if not servicio:
        servicio = models.Servicio(codigo=nombre[:20], nombre=nombre, estado=True)
        db.add(servicio)
        db.flush()
    return servicio


def get_or_create_tecnologo(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    if not nombre:
        return None
    t = db.query(models.Tecnologo).filter_by(nombre=nombre).first()
    if not t:
        t = models.Tecnologo(nombre=nombre, estado=True)
        db.add(t)
        db.flush()
    return t


def get_or_create_radiologo(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    if not nombre:
        return None
    r = db.query(models.Radiologo).filter_by(nombre=nombre).first()
    if not r:
        r = models.Radiologo(nombre=nombre, estado=True)
        db.add(r)
        db.flush()
    return r


def get_or_create_transcriptora(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    if not nombre:
        return None
    t = db.query(models.Transcriptora).filter_by(nombre=nombre).first()
    if not t:
        t = models.Transcriptora(nombre=nombre, estado=True)
        db.add(t)
        db.flush()
    return t


def get_or_create_reporte(db, nombre_raw):
    nombre = normalize_name(nombre_raw)
    if not nombre:
        return None
    r = db.query(models.Reporte).filter_by(nombre=nombre).first()
    if not r:
        r = models.Reporte(nombre=nombre, estado=True)
        db.add(r)
        db.flush()
    return r


def parse_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def parse_float(val):
    try:
        f = float(str(val).replace(",", ".").strip())
        return f if f > 0 else 0.01
    except (ValueError, TypeError):
        return 0.01


def parse_int(val):
    try:
        i = int(float(str(val).strip()))
        return i if i > 0 else 1
    except (ValueError, TypeError):
        return 1


_counters = {}


def generate_codigo(year, month):
    key = (year, month)
    _counters[key] = _counters.get(key, 0) + 1
    return f"RX-{year}-{month:02d}-{_counters[key]:06d}"


def import_excel(filepath: str):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecute: pip install openpyxl")
        return

    print(f"Abriendo: {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    print(f"Hoja activa: '{ws.title}', {ws.max_row - 1} filas de datos")

    db = SessionLocal()
    try:
        seed_initial_data(db)
        na_causa = db.query(models.CausaRechazo).filter_by(descripcion="NA").first()

        imported = 0
        skipped = 0
        errors = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):
                    continue

                fecha_toma = parse_datetime(row[0])
                fecha_orden = parse_datetime(row[1])
                identificacion = str(row[2]).strip() if row[2] else None
                nombre_paciente = normalize_name(row[3]) if row[3] else None

                if not fecha_toma or not identificacion or not nombre_paciente:
                    skipped += 1
                    continue

                if fecha_orden is None or fecha_orden > fecha_toma:
                    fecha_orden = fecha_toma

                estudio = get_or_create_estudio(db, row[4])
                servicio = get_or_create_servicio(db, row[5])
                numero_tomas = parse_int(row[6])
                kv = parse_float(row[7])
                mas = parse_float(row[8])

                rechazada_raw = normalize_name(row[9]) if row[9] else "NO"
                rechazada = rechazada_raw == "SI"

                causa_raw = normalize_name(row[10]) if row[10] else "NA"
                if rechazada and causa_raw and causa_raw != "NA":
                    causa = db.query(models.CausaRechazo).filter_by(descripcion=causa_raw).first()
                    if not causa:
                        causa = models.CausaRechazo(descripcion=causa_raw, estado=True)
                        db.add(causa)
                        db.flush()
                else:
                    causa = na_causa

                tecnologo = get_or_create_tecnologo(db, row[11])
                reporte = get_or_create_reporte(db, row[12])

                fecha_lectura = parse_datetime(row[13])
                radiologo = get_or_create_radiologo(db, row[14])
                transcriptora = get_or_create_transcriptora(db, row[15])
                observaciones = str(row[16]).strip() if row[16] else None

                estado = "LEIDO" if radiologo else "PENDIENTE"
                if rechazada:
                    estado = "LEIDO" if radiologo else "PENDIENTE"

                codigo = generate_codigo(fecha_toma.year, fecha_toma.month)

                registro = models.RegistroRayosX(
                    codigo_registro=codigo,
                    fecha_hora_toma=fecha_toma,
                    fecha_hora_orden=fecha_orden,
                    identificacion=identificacion,
                    nombre_paciente=nombre_paciente,
                    estudio_id=estudio.id if estudio else None,
                    servicio_id=servicio.id if servicio else None,
                    numero_tomas=numero_tomas,
                    kv=kv,
                    mas=mas,
                    rechazada=rechazada,
                    causa_rechazo_id=causa.id if causa else None,
                    tecnologo_id=tecnologo.id if tecnologo else None,
                    reporte_id=reporte.id if reporte else None,
                    fecha_hora_lectura_radiologo=fecha_lectura,
                    radiologo_id=radiologo.id if radiologo else None,
                    transcriptora_id=transcriptora.id if transcriptora else None,
                    observaciones=observaciones,
                    estado=estado,
                )
                db.add(registro)
                db.commit()

                if imported % 100 == 0:
                    print(f"  Procesadas {row_num - 1} filas... ({imported} importadas)")

                imported += 1

            except Exception as e:
                db.rollback()
                print(f"  ERROR en fila {row_num}: {type(e).__name__}: {str(e)[:120]}")
                errors += 1
                continue
        print(f"\nImportacion completada OK:")
        print(f"  Importados: {imported}")
        print(f"  Omitidos (datos incompletos): {skipped}")
        print(f"  Errores: {errors}")

    except Exception as e:
        db.rollback()
        print(f"ERROR CRÍTICO: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Default path relative to script location
        default = Path(__file__).parent.parent / "Informe de rayos X CMVA.xlsx"
        if default.exists():
            import_excel(str(default))
        else:
            print("Uso: python import_excel.py <ruta_excel>")
            print(f"No se encontró el archivo en: {default}")
            sys.exit(1)
    else:
        import_excel(sys.argv[1])
